import system
import openpyxl as ox
import numpy as np
import pickle

wb = ox.load_workbook("data/exp_curve.xlsx")
util_sheet = wb.get_sheet_by_name('utilities')

erg = dict()

for flex in np.arange(0,1.1,0.1):
#for flex in np.arange(0,0.1,0.1):
	util_sheet["C15"].value = flex

	for cell, in util_sheet["C2:C6"]:
		print(util_sheet.cell(row=19+cell.row-2,column=3).value)
		cell.value = util_sheet.cell(row=19+cell.row-2,column=3).value * (1- util_sheet["C15"].value )

	for cell, in util_sheet["B2:B6"]:

		cell.value = util_sheet.cell(row=cell.row,column=3).value * (1- util_sheet["C16"].value )

	for cell, in util_sheet["E2:E6"]:

		cell.value = util_sheet.cell(row=19+cell.row-2,column=3).value * util_sheet["C15"].value

	for cell, in util_sheet["D2:D6"]:

		cell.value = util_sheet.cell(row=cell.row,column=5).value * (1- util_sheet["C17"].value )

	wb.save("run.xlsx")


	sys = system.System.loadFromXLSX('run.xlsx')
	system.LOG_DEMAND = True
	system._USE_GRAPH_MATCH = False
	for u in sys.users:
		if u.isRemote:
			u._isRemote = False

	# AddWind
	#max_idx = max([p.idx for p in sys.producers]) + 1
	#for idx in range(1):
	#	sys.producers.append(system.WindAgent(sys,max_idx+idx,2000,10))


	sys.runForDays(1)
	sys.resetCosts()
	sys.runForDays(1)

	erg[flex] = sys.costPerMW()


with open("rem_match_.pkl", 'wb') as f:
	pickle.dump(erg, f, pickle.HIGHEST_PROTOCOL)